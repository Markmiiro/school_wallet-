# ================================================
# app/routes/analytics.py
# ------------------------------------------------
# Phase 6 — Analytics Dashboard
#
# ENDPOINTS:
# GET /analytics/school/{id}/overview    → school summary
# GET /analytics/school/{id}/daily       → daily breakdown
# GET /analytics/school/{id}/weekly      → weekly trends
# GET /analytics/school/{id}/classes     → class level spending
# GET /analytics/student/{id}/summary    → per student summary
# GET /analytics/school/{id}/export      → Excel export
# ================================================

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, date, timedelta
from typing import Optional
import io

from app.database import get_db
from app.models import (
    Transaction, Wallet, Student,
    Merchant, School, User
)

router = APIRouter()


# ================================================
# HELPER FUNCTIONS
# ================================================

def get_transactions_for_school(db: Session, school_id: int, target_date: date):
    """Get all completed payment transactions for a school on a given date."""
    # Get all merchants for this school
    merchants = db.query(Merchant).filter(
        Merchant.school_id == school_id
    ).all()

    merchant_ids = [m.id for m in merchants]

    if not merchant_ids:
        return []

    # Get all transactions for these merchants
    txns = (
        db.query(Transaction)
        .filter(
            Transaction.merchant_id.in_(merchant_ids),
            Transaction.type == "payment",
            Transaction.status == "completed",
        )
        .all()
    )

    # Filter by date
    return [
        t for t in txns
        if t.timestamp and t.timestamp.date() == target_date
    ]


def get_student_name(db: Session, wallet_id: int) -> str:
    """Get student name from wallet ID."""
    wallet = db.query(Wallet).filter(Wallet.id == wallet_id).first()
    if not wallet:
        return "Unknown"
    student = db.query(Student).filter(
        Student.id == wallet.student_id
    ).first()
    return student.name if student else "Unknown"


# ================================================
# ENDPOINT 1 — School overview
# ================================================
@router.get("/school/{school_id}/overview")
def school_overview(
    school_id: int,
    db: Session = Depends(get_db)
):
    """
    High level overview for headteacher.
    Shows today, this week, this month at a glance.
    Perfect for the headteacher's morning briefing.
    """
    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")

    today       = date.today()
    week_start  = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    # Get all merchants
    merchants = db.query(Merchant).filter(
        Merchant.school_id == school_id
    ).all()
    merchant_ids = [m.id for m in merchants]

    if not merchant_ids:
        return {
            "school": school.name,
            "message": "No merchants registered yet"
        }

    # Get ALL completed transactions
    all_txns = (
        db.query(Transaction)
        .filter(
            Transaction.merchant_id.in_(merchant_ids),
            Transaction.type == "payment",
            Transaction.status == "completed",
        )
        .all()
    )

    # Filter by period
    today_txns  = [t for t in all_txns if t.timestamp and t.timestamp.date() == today]
    week_txns   = [t for t in all_txns if t.timestamp and t.timestamp.date() >= week_start]
    month_txns  = [t for t in all_txns if t.timestamp and t.timestamp.date() >= month_start]

    # Get all students in school
    total_students = db.query(Student).filter(
        Student.school_id == school_id
    ).count()

    # Students who spent today
    today_wallet_ids = set(t.wallet_id for t in today_txns)
    active_students_today = len(today_wallet_ids)

    # Top spending student today
    student_totals = {}
    for t in today_txns:
        student_totals[t.wallet_id] = student_totals.get(t.wallet_id, 0) + t.amount

    top_student_today = None
    if student_totals:
        top_wallet_id = max(student_totals, key=student_totals.get)
        top_student_today = {
            "name":       get_student_name(db, top_wallet_id),
            "amount_ugx": student_totals[top_wallet_id]
        }

    # Top merchant today
    merchant_totals = {}
    for t in today_txns:
        merchant_totals[t.merchant_id] = merchant_totals.get(t.merchant_id, 0) + t.amount

    top_merchant_today = None
    if merchant_totals:
        top_merchant_id    = max(merchant_totals, key=merchant_totals.get)
        top_merchant_obj   = db.query(Merchant).filter(
            Merchant.id == top_merchant_id
        ).first()
        top_merchant_today = {
            "name":       top_merchant_obj.name if top_merchant_obj else "Unknown",
            "amount_ugx": merchant_totals[top_merchant_id]
        }

    return {
        "school":          school.name,
        "school_id":       school_id,
        "as_of":           str(today),
        "students": {
            "total_enrolled":     total_students,
            "active_today":       active_students_today,
            "inactive_today":     total_students - active_students_today,
        },
        "today": {
            "total_spent_ugx":    sum(t.amount for t in today_txns),
            "num_transactions":   len(today_txns),
            "top_student":        top_student_today,
            "top_merchant":       top_merchant_today,
        },
        "this_week": {
            "total_spent_ugx":    sum(t.amount for t in week_txns),
            "num_transactions":   len(week_txns),
            "daily_average_ugx":  round(sum(t.amount for t in week_txns) / 7),
        },
        "this_month": {
            "total_spent_ugx":    sum(t.amount for t in month_txns),
            "num_transactions":   len(month_txns),
        },
        "merchants": {
            "total":   len(merchants),
            "names":   [m.name for m in merchants],
        }
    }


# ================================================
# ENDPOINT 2 — Daily breakdown
# ================================================
@router.get("/school/{school_id}/daily")
def school_daily_breakdown(
    school_id: int,
    report_date: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    """
    Hour by hour breakdown of spending for a school.
    Shows when students spend most during the day.
    """
    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")

    if report_date:
        try:
            target_date = datetime.strptime(report_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Use YYYY-MM-DD format")
    else:
        target_date = date.today()

    txns = get_transactions_for_school(db, school_id, target_date)

    # Group by hour
    hourly = {}
    for hour in range(6, 20):  # 6am to 8pm
        hourly[hour] = {"total_ugx": 0, "num_transactions": 0}

    for t in txns:
        if t.timestamp:
            hour = t.timestamp.hour
            if hour in hourly:
                hourly[hour]["total_ugx"]        += t.amount
                hourly[hour]["num_transactions"] += 1

    # Format hourly breakdown
    hourly_list = []
    for hour, data in hourly.items():
        am_pm = "AM" if hour < 12 else "PM"
        display_hour = hour if hour <= 12 else hour - 12
        hourly_list.append({
            "hour":             f"{display_hour}:00 {am_pm}",
            "total_ugx":        data["total_ugx"],
            "num_transactions": data["num_transactions"],
        })

    # Peak spending time
    peak = max(hourly_list, key=lambda x: x["total_ugx"])

    return {
        "school":      school.name,
        "date":        str(target_date),
        "total_ugx":   sum(t.amount for t in txns),
        "peak_hour":   peak["hour"] if peak["total_ugx"] > 0 else "No transactions",
        "hourly":      hourly_list,
    }


# ================================================
# ENDPOINT 3 — Weekly trends
# ================================================
@router.get("/school/{school_id}/weekly")
def school_weekly_trends(
    school_id: int,
    db: Session = Depends(get_db)
):
    """
    7-day spending trend for a school.
    Shows which days students spend most.
    """
    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")

    today      = date.today()
    week_start = today - timedelta(days=6)  # last 7 days

    merchants    = db.query(Merchant).filter(Merchant.school_id == school_id).all()
    merchant_ids = [m.id for m in merchants]

    if not merchant_ids:
        return {"school": school.name, "message": "No merchants found"}

    all_txns = (
        db.query(Transaction)
        .filter(
            Transaction.merchant_id.in_(merchant_ids),
            Transaction.type == "payment",
            Transaction.status == "completed",
        )
        .all()
    )

    # Build 7-day breakdown
    weekly = []
    total_7_days = 0

    for i in range(7):
        day      = week_start + timedelta(days=i)
        day_txns = [t for t in all_txns if t.timestamp and t.timestamp.date() == day]
        day_total = sum(t.amount for t in day_txns)
        total_7_days += day_total

        weekly.append({
            "date":             str(day),
            "day_name":         day.strftime("%A"),
            "total_ugx":        day_total,
            "num_transactions": len(day_txns),
            "is_today":         day == today,
        })

    # Best and worst day
    best_day  = max(weekly, key=lambda x: x["total_ugx"])
    worst_day = min(weekly, key=lambda x: x["total_ugx"])

    return {
        "school":           school.name,
        "period":           f"{week_start} to {today}",
        "total_7_days_ugx": total_7_days,
        "daily_average_ugx": round(total_7_days / 7),
        "best_day":         best_day["day_name"],
        "worst_day":        worst_day["day_name"],
        "days":             weekly,
    }


# ================================================
# ENDPOINT 4 — Student spending summary
# ================================================
@router.get("/student/{student_id}/summary")
def student_spending_summary(
    student_id: int,
    db: Session = Depends(get_db)
):
    """
    Full spending summary for one student.
    Useful for parents to review their child's spending.
    """
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    wallet = db.query(Wallet).filter(
        Wallet.student_id == student_id
    ).first()
    if not wallet:
        raise HTTPException(status_code=404, detail="Wallet not found")

    today       = date.today()
    week_start  = today - timedelta(days=6)
    month_start = today.replace(day=1)

    all_txns = (
        db.query(Transaction)
        .filter(
            Transaction.wallet_id == wallet.id,
            Transaction.type == "payment",
            Transaction.status == "completed",
        )
        .order_by(Transaction.timestamp.desc())
        .all()
    )

    today_txns  = [t for t in all_txns if t.timestamp and t.timestamp.date() == today]
    week_txns   = [t for t in all_txns if t.timestamp and t.timestamp.date() >= week_start]
    month_txns  = [t for t in all_txns if t.timestamp and t.timestamp.date() >= month_start]

    # Most visited merchant
    merchant_visits = {}
    for t in all_txns:
        if t.merchant_id:
            merchant_visits[t.merchant_id] = merchant_visits.get(t.merchant_id, 0) + 1

    favourite_merchant = None
    if merchant_visits:
        top_merchant_id  = max(merchant_visits, key=merchant_visits.get)
        top_merchant_obj = db.query(Merchant).filter(
            Merchant.id == top_merchant_id
        ).first()
        favourite_merchant = top_merchant_obj.name if top_merchant_obj else "Unknown"

    # Recent transactions
    recent = []
    for t in all_txns[:10]:
        merchant_obj = db.query(Merchant).filter(
            Merchant.id == t.merchant_id
        ).first()
        recent.append({
            "date":        t.timestamp.strftime("%d %b %Y %I:%M %p") if t.timestamp else "N/A",
            "merchant":    merchant_obj.name if merchant_obj else "Unknown",
            "amount_ugx":  t.amount,
            "description": t.description or "Payment",
        })

    return {
        "student":          student.name,
        "student_id":       student_id,
        "current_balance":  wallet.balance,
        "daily_limit":      wallet.daily_limit,
        "currency":         "UGX",
        "favourite_merchant": favourite_merchant,
        "spending": {
            "today_ugx":    sum(t.amount for t in today_txns),
            "this_week_ugx": sum(t.amount for t in week_txns),
            "this_month_ugx": sum(t.amount for t in month_txns),
            "all_time_ugx": sum(t.amount for t in all_txns),
        },
        "transactions": {
            "today":     len(today_txns),
            "this_week": len(week_txns),
            "this_month": len(month_txns),
            "all_time":  len(all_txns),
        },
        "recent_transactions": recent,
    }


# ================================================
# ENDPOINT 5 — Export to Excel
# ================================================
@router.get("/school/{school_id}/export")
def export_to_excel(
    school_id: int,
    report_date: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    """
    Export school transactions to Excel.
    Downloads a .xlsx file for the bursar.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Add it to requirements.txt"
        )

    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")

    if report_date:
        try:
            target_date = datetime.strptime(report_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Use YYYY-MM-DD format")
    else:
        target_date = date.today()

    txns = get_transactions_for_school(db, school_id, target_date)

    # ── Build Excel workbook ─────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Transactions {target_date}"

    # Header style
    header_font    = Font(bold=True, color="FFFFFF")
    header_fill    = PatternFill("solid", fgColor="1B5E20")
    header_align   = Alignment(horizontal="center")

    # Headers
    headers = [
        "Time", "Student", "Merchant",
        "Amount (UGX)", "Description", "Transaction ID"
    ]

    for col, header in enumerate(headers, 1):
        cell               = ws.cell(row=1, column=col, value=header)
        cell.font          = header_font
        cell.fill          = header_fill
        cell.alignment     = header_align

    # Data rows
    total = 0
    for row, txn in enumerate(txns, 2):
        student_name = get_student_name(db, txn.wallet_id)
        merchant_obj = db.query(Merchant).filter(
            Merchant.id == txn.merchant_id
        ).first()
        merchant_name = merchant_obj.name if merchant_obj else "Unknown"

        ws.cell(row=row, column=1, value=txn.timestamp.strftime("%I:%M %p") if txn.timestamp else "N/A")
        ws.cell(row=row, column=2, value=student_name)
        ws.cell(row=row, column=3, value=merchant_name)
        ws.cell(row=row, column=4, value=txn.amount)
        ws.cell(row=row, column=5, value=txn.description or "Payment")
        ws.cell(row=row, column=6, value=txn.id)

        total += txn.amount

    # Total row
    total_row = len(txns) + 2
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=total).font   = Font(bold=True)

    # Auto-size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"school_wallet_{school.name}_{target_date}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )